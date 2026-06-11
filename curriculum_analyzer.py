# curriculum_analyzer.py
import os
import re
import csv
import pandas as pd
import openpyxl
from collections import Counter
import difflib
import time
from mistralai.client import Mistral

# Standard Chapter Codes for Grade 6 Science
CHAPTER_CODES = {
    "materials around us": "MAT",
    "living creatures": "LIV",
    "living creatures: exploring their characteristics": "LIV",
    "living creatures: exploring\ntheir characteristics": "LIV",
    "the wonderful world of science": "SCI",
    "diversity in the living world": "DIV",
    "exploring magnets": "MAG",
    "mindful eating": "EAT",
    "mindful eating: a path  to a healthy body": "EAT",
    "mindful eating: a path to a healthy body": "EAT",
    "measurement of length and motion": "MEA",
    "measurement of length\nand motion": "MEA",
    "temperature and its measurement": "TEM",
    "a journey through states of water": "WAT",
    "methods of separation": "SEP",
    "methods of separation in everyday life": "SEP",
    "method of separation in everyday life": "SEP",
    "nature's treasures": "NAT",
    "nature’s treasures": "NAT",
    "beyond earth": "BEY"
}

# Predefined Topic Codes for Grade 6 Science
TOPIC_CODES = {
    "observing objects around us": "OBS",
    "how to group materials?": "GRP",
    "what are the different properties of materials?": "PROP",
    "what are the different properties of\nmaterials?\n": "PROP",
    "what is matter?": "MATT",
    
    "magnetic and non-magnetic materials": "MAT",
    "poles of a magnet": "POLE",
    "finding directions": "DIR",
    "interaction between magnets": "INT",
    "properties of magnets": "PROP",
    
    "what sets the living apart from the non-living?": "LIV",
    "characteristics of living things": "CHAR",
    "seed germination": "GERM",
    "life cycle of a plant": "PLNT",
    "life cycle of animals": "ANIM",
}

def normalize_text(text):
    if not text:
        return ""
    text = str(text).lower().strip()
    text = re.sub(r"[\s\n\r\t]+", " ", text)
    text = text.strip("?.;,")
    return text

def get_next_sequence_number(chapter_code, topic_code, output_dir="output"):
    return 1

def find_in_blueprint(topic_query):
    blueprint_path = r"1. Test Blueprint _Grade 6 (1).xlsx"
    if not os.path.exists(blueprint_path):
        blueprint_path = "1. Test Blueprint _Grade 6 (1).xlsx"
    if not os.path.exists(blueprint_path):
        return None
        
    try:
        wb = openpyxl.load_workbook(blueprint_path, data_only=True)
        norm_query = normalize_text(topic_query)
        
        for sheet_name in wb.sheetnames:
            if sheet_name == "Summary":
                continue
            sheet = wb[sheet_name]
            current_topic = None
            for r_idx in range(4, sheet.max_row + 1):
                t_val = sheet.cell(row=r_idx, column=1).value
                if t_val:
                    current_topic = str(t_val).strip()
                
                if current_topic and normalize_text(current_topic) == norm_query:
                    lo_val = sheet.cell(row=r_idx, column=2).value
                    if lo_val:
                        return {
                            "chapter_name": sheet_name.strip(),
                            "topic_name": current_topic,
                            "learning_outcome": str(lo_val).strip()
                        }
    except Exception as e:
        print(f"Error reading blueprint: {e}")
    return None

def load_curriculum_framework(framework_path="Final_Class 6_Science_Mastery Framework_ (1).xlsx"):
    if not os.path.exists(framework_path):
        parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if '__file__' in locals() else os.path.dirname(os.getcwd())
        framework_path = os.path.join(parent_dir, "Final_Class 6_Science_Mastery Framework_ (1).xlsx")
    if not os.path.exists(framework_path):
        framework_path = os.path.join(os.path.dirname(os.getcwd()), "Final_Class 6_Science_Mastery Framework_ (1).xlsx")
    if not os.path.exists(framework_path):
        framework_path = "Final_Class 6_Science_Mastery Framework_ (1).xlsx"
        
    if not os.path.exists(framework_path):
        return []
    try:
        wb = openpyxl.load_workbook(framework_path, data_only=True)
        sheet = wb["Framework"]
        
        # Read headers
        headers = []
        for col_idx in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col_idx).value
            headers.append(str(val).strip() if val is not None else "")
            
        def find_index(names):
            for name in names:
                for idx, h in enumerate(headers):
                    if h.lower() == name.lower():
                        return idx
            return -1

        cg_idx = find_index(["CG #", "NCF CG #", "Curricular Goal #"])
        comp_idx = find_index(["Competency", "Comp #"])
        chap_idx = find_index(["Grade 6 Chapter", "Chapter"])
        topic_idx = find_index(["Grade 6 Topic/ Sub- topic", "Topic/ Sub- topic", "Topic"])
        lo_idx = find_index(["Learning Outcome", "LO"])

        if cg_idx == -1: cg_idx = 0
        if comp_idx == -1: comp_idx = 2
        if chap_idx == -1: chap_idx = 4
        if topic_idx == -1: topic_idx = 5
        if lo_idx == -1: lo_idx = 6

        unique_combos = []
        seen = set()
        
        for r_idx in range(2, sheet.max_row + 1):
            cg = str(sheet.cell(row=r_idx, column=cg_idx + 1).value or "").strip()
            comp = str(sheet.cell(row=r_idx, column=comp_idx + 1).value or "").strip()
            chap = str(sheet.cell(row=r_idx, column=chap_idx + 1).value or "").strip()
            topic = str(sheet.cell(row=r_idx, column=topic_idx + 1).value or "").strip()
            lo = str(sheet.cell(row=r_idx, column=lo_idx + 1).value or "").strip()
            
            cg = re.sub(r"[\s\n\r\t]+", " ", cg).strip()
            comp = re.sub(r"[\s\n\r\t]+", " ", comp).strip()
            chap = re.sub(r"[\s\n\r\t]+", " ", chap).strip()
            topic = re.sub(r"[\s\n\r\t]+", " ", topic).strip()
            lo = re.sub(r"[\s\n\r\t]+", " ", lo).strip()
            
            if not (cg or comp or chap or topic or lo):
                continue
                
            key = (cg, comp, chap, topic, lo)
            if key not in seen:
                seen.add(key)
                unique_combos.append({
                    "ncf_cg": cg,
                    "competency": comp,
                    "chapter_name": chap,
                    "topic_name": topic,
                    "learning_outcome": lo
                })
        return unique_combos
    except Exception as e:
        print(f"Error loading framework: {e}")
        return []

def analyze_curriculum_alignment(uploaded_file, api_key: str):
    """
    Analyzes the uploaded question bank spreadsheet and queries the framework/blueprint.
    Returns a dictionary of alignment fields and starting sequence number.
    """
    framework_path = r"Final_Class 6_Science_Mastery Framework_ (1).xlsx"
    if not os.path.exists(framework_path):
        parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if '__file__' in locals() else os.path.dirname(os.getcwd())
        framework_path = os.path.join(parent_dir, "Final_Class 6_Science_Mastery Framework_ (1).xlsx")
    if not os.path.exists(framework_path):
        framework_path = os.path.join(os.path.dirname(os.getcwd()), "Final_Class 6_Science_Mastery Framework_ (1).xlsx")
    if not os.path.exists(framework_path):
        framework_path = "Final_Class 6_Science_Mastery Framework_ (1).xlsx"
        
    # Default fallback values
    result = {
        "chapter_name": "Materials Around Us",
        "chapter_code": "MAT",
        "topic_code": "PROP",
        "ncf_cg": "CG-1",
        "competency": "C-1.1",
        "learning_outcome": "Identifies and explains different properties of materials and relates them to their uses.",
        "lo_id": "LO-1.1.a",
        "start_seq": 1
    }
    
    try:
        # 1. Read sheet Questions from uploaded file
        uploaded_file.seek(0)
        wb_upload = openpyxl.load_workbook(uploaded_file, data_only=True)
        matching_sheet = None
        for name in wb_upload.sheetnames:
            if name.strip().lower() == "questions":
                matching_sheet = name
                break
        if matching_sheet is None and len(wb_upload.sheetnames) > 0:
            matching_sheet = wb_upload.sheetnames[0]
        if matching_sheet is None:
            return result
        sheet = wb_upload[matching_sheet]
        
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        topic_idx = -1
        stem_idx = -1
        for idx, h in enumerate(headers):
            if h.lower() == "topic":
                topic_idx = idx + 1
            elif h.lower() == "item_stem":
                stem_idx = idx + 1
                
        if topic_idx == -1:
            return result
            
        topics = []
        sample_stems = []
        for r_idx in range(2, sheet.max_row + 1):
            val = sheet.cell(row=r_idx, column=topic_idx).value
            stem_val = sheet.cell(row=r_idx, column=stem_idx).value if stem_idx != -1 else None
            if val:
                topics.append(str(val).strip())
            if stem_val:
                sample_stems.append(str(stem_val).strip())
                
        if not topics:
            return result
            
        most_common_topic = Counter(topics).most_common(1)[0][0]
        norm_query = normalize_text(most_common_topic)
        
        # Load all framework entries dynamically
        unique_combos = load_curriculum_framework(framework_path)
        
        matched_idx = -1
        
        # 2. Heuristic Matching (used as default / fallback)
        # Try exact topic match
        for idx, combo in enumerate(unique_combos):
            if normalize_text(combo["topic_name"]) == norm_query:
                matched_idx = idx
                break
                
        # Try fuzzy topic match
        if matched_idx == -1:
            best_ratio = 0.0
            for idx, combo in enumerate(unique_combos):
                ratio = difflib.SequenceMatcher(None, normalize_text(combo["topic_name"]), norm_query).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    matched_idx = idx
            if best_ratio < 0.7:
                matched_idx = -1
                
        # Try blueprint match if topic not found in framework
        if matched_idx == -1:
            bp_match = find_in_blueprint(most_common_topic)
            if bp_match:
                norm_lo = normalize_text(bp_match["learning_outcome"])
                for idx, combo in enumerate(unique_combos):
                    if normalize_text(combo["learning_outcome"]) == norm_lo:
                        matched_idx = idx
                        break
                        
        if matched_idx == -1 and unique_combos:
            matched_idx = 0
            
        # Apply the matched framework entry (as base configuration)
        if matched_idx != -1 and matched_idx < len(unique_combos):
            matched_combo = unique_combos[matched_idx]
            result["chapter_name"] = matched_combo["chapter_name"]
            result["ncf_cg"] = matched_combo["ncf_cg"]
            result["competency"] = matched_combo["competency"]
            result["learning_outcome"] = matched_combo["learning_outcome"]
            
            # Map Chapter Code to Chapter ID
            chapter_code = None
            science_path = "Science Class 6.xlsx"
            if not os.path.exists(science_path):
                science_path = os.path.join(os.path.dirname(os.getcwd()), "Science Class 6.xlsx")
            if not os.path.exists(science_path):
                parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if '__file__' in locals() else os.path.dirname(os.getcwd())
                science_path = os.path.join(parent_dir, "Science Class 6.xlsx")
                
            try:
                df_sci = pd.read_excel(science_path)
                norm_chap_target = re.sub(r'\s+', ' ', result["chapter_name"].lower().strip())
                for idx, r in df_sci.iterrows():
                    chap_val = r.get("Chapter")
                    cid_val = r.get("Chapter ID")
                    if chap_val and cid_val:
                        norm_chap_val = re.sub(r'\s+', ' ', str(chap_val).lower().strip())
                        if norm_chap_val == norm_chap_target:
                            chapter_code = str(cid_val).strip()
                            break
            except Exception:
                pass
                
            if not chapter_code:
                # Fallback mapping
                chapter_code = {
                    "a journey through states of water": "67ff3d8b2613c0dfa772579b",
                    "beyond earth": "67ff3d8a2613c0dfa7725785",
                    "diversity in the living world": "67ff3d8a2613c0dfa7725742",
                    "exploring magnets": "67ff3d8a2613c0dfa772574f",
                    "living creatures": "67ff3d8b2613c0dfa7725792",
                    "living creatures: exploring their characteristics": "67ff3d8b2613c0dfa7725792",
                    "materials around us": "67ff3d8a2613c0dfa7725765",
                    "measurement of length and motion": "67ff3d8a2613c0dfa772576c",
                    "method of separation in everyday life": "68d260c812d191aeb6c953e6",
                    "methods of separation": "68d260c812d191aeb6c953e6",
                    "mindful eating": "67ff3d8a2613c0dfa772575a",
                    "mindful eating: a path to a healthy body": "67ff3d8a2613c0dfa772575a",
                    "nature's treasures": "67ff3d8b2613c0dfa77257af",
                    "nature’s treasures": "67ff3d8b2613c0dfa77257af",
                    "temperature and its measurement": "67ff3d8a2613c0dfa7725771",
                    "the wonderful world of science": "68d2604a12d191aeb6c9320b",
                }.get(re.sub(r'\s+', ' ', result["chapter_name"].lower().strip()))
                
            if not chapter_code:
                chapter_code = result["chapter_name"][:3].upper()
            result["chapter_code"] = chapter_code
            
            # Map Topic Code
            norm_topic = normalize_text(most_common_topic)
            topic_code = None
            for key, val in TOPIC_CODES.items():
                if normalize_text(key) == norm_topic:
                    topic_code = val
                    break
            if not topic_code:
                topic_code = "".join([w[0].upper() for w in norm_topic.split() if w])[:4]
            result["topic_code"] = topic_code
            
            # Generate LO ID
            comp_num = "1.1"
            m_comp = re.search(r"C-(\d+\.\d+)", result["competency"])
            if m_comp:
                comp_num = m_comp.group(1)
            else:
                m_num = re.search(r"(\d+\.\d+)", result["competency"])
                if m_num:
                    comp_num = m_num.group(1)
            
            comp_los = []
            seen_lo = set()
            for combo in unique_combos:
                if combo["competency"] == result["competency"]:
                    lo_clean = combo["learning_outcome"].strip()
                    if lo_clean not in seen_lo:
                        seen_lo.add(lo_clean)
                        comp_los.append(lo_clean)
            
            try:
                curr_lo = result["learning_outcome"].strip()
                match_lo_idx = 0
                for idx, lo in enumerate(comp_los):
                    if normalize_text(lo) == normalize_text(curr_lo):
                        match_lo_idx = idx
                        break
                suffix_letter = chr(ord('a') + match_lo_idx)
            except Exception:
                suffix_letter = 'a'
                
            result["lo_id"] = f"LO-{comp_num}.{suffix_letter}"

        # 3. AI-Assisted Match (if API Key is not mock/empty)
        if api_key and api_key.lower() not in ("mock", "test") and unique_combos:
            try:
                # Format unique combos for the prompt
                combos_text = ""
                for idx, combo in enumerate(unique_combos):
                    combos_text += (
                        f"Index {idx}:\n"
                        f"  Chapter: {combo['chapter_name']}\n"
                        f"  Topic: {combo['topic_name']}\n"
                        f"  Learning Outcome: {combo['learning_outcome']}\n"
                        f"  NCF CG #: {combo['ncf_cg']}\n"
                        f"  Competency: {combo['competency']}\n\n"
                    )

                # Sample stems text
                sample_stems_text = ""
                for s_idx, stem in enumerate(sample_stems[:3], start=1):
                    sample_stems_text += f"{s_idx}. {stem}\n"

                client = Mistral(api_key=api_key)
                system_prompt = (
                    "You are a curriculum mapping expert. Your task is to match an assessment topic and its sample questions "
                    "to the correct Indian NCF 2023 Grade 6 Science curriculum framework entry, and suggest code mappings.\n\n"
                    "UPLOADING DETAILS:\n"
                    f"Uploaded Topic: {most_common_topic}\n"
                    "Sample Question Stems:\n"
                    f"{sample_stems_text}\n\n"
                    "LIST OF FRAMEWORK ENTRIES:\n"
                    f"{combos_text}\n"
                    "INSTRUCTIONS:\n"
                    "1. Find the entry index in the framework list that best matches the concept and topic being assessed in the questions.\n"
                    "2. Suggest a 3-4 letter uppercase Chapter Code (e.g. 'MAT' for 'Materials Around Us', 'LIV' for 'Living Creatures').\n"
                    "3. Suggest a 3-4 letter uppercase Topic Code (e.g. 'PROP' for 'Properties of Materials', 'OBS' for 'Observing Objects').\n"
                    "4. Determine the LO ID in the format LO-[Competency_Number].[letter] (e.g. LO-1.1.a). The suffix letter ('a', 'b', 'c', etc.) must correspond to the order of this learning outcome under its competency.\n\n"
                    "OUTPUT format must be JSON only:\n"
                    "{\n"
                    "  \"matched_index\": <integer index of matched framework entry>,\n"
                    "  \"chapter_code\": \"<uppercase code>\",\n"
                    "  \"topic_code\": \"<uppercase code>\",\n"
                    "  \"lo_id\": \"<suggested LO ID>\"\n"
                    "}"
                )
                
                response = client.chat.complete(
                    model="mistral-large-latest",
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": "Please suggest matched framework entry index and codes."}
                    ],
                    temperature=0.1,
                    response_format={"type": "json_object"}
                )
                
                content = response.choices[0].message.content.strip()
                import json
                cleaned = content
                if cleaned.startswith("```"):
                    lines = cleaned.split("\n")
                    if lines[0].startswith("```"):
                        lines = lines[1:]
                    if lines[-1].startswith("```"):
                        lines = lines[:-1]
                    cleaned = "\n".join(lines).strip()
                ai_res = json.loads(cleaned)
                
                ai_idx = ai_res.get("matched_index")
                if ai_idx is not None and isinstance(ai_idx, int) and 0 <= ai_idx < len(unique_combos):
                    matched_combo = unique_combos[ai_idx]
                    result["chapter_name"] = matched_combo["chapter_name"]
                    result["ncf_cg"] = matched_combo["ncf_cg"]
                    result["competency"] = matched_combo["competency"]
                    result["learning_outcome"] = matched_combo["learning_outcome"]
                
                if ai_res.get("chapter_code"):
                    result["chapter_code"] = str(ai_res["chapter_code"]).strip().upper()
                if ai_res.get("topic_code"):
                    result["topic_code"] = str(ai_res["topic_code"]).strip().upper()
                if ai_res.get("lo_id"):
                    result["lo_id"] = str(ai_res["lo_id"]).strip()
            except Exception as exc:
                print(f"AI alignment mapping failed, using heuristics: {exc}")
                
        # 4. Resolve Starting Sequence Number
        result["start_seq"] = get_next_sequence_number(result["chapter_code"], result["topic_code"])
        
    except Exception as exc:
        print(f"Error during curriculum alignment analysis: {exc}")
        
    return result
